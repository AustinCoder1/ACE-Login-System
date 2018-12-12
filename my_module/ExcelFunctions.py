#!/usr/bin/env python
# coding: utf-8

# In[1]:


from xlwt import Workbook
from xlrd import open_workbook
from pathlib import Path
from tkinter import messagebox
import openpyxl
import string
from my_module.Security import *

class excelFunctions():
    def __init__(self):
        if (os.path.isfile("login.xlsx")):
            self.rowCount = 0
            self.columnNumber = 0
            self.loadBook = openpyxl.load_workbook("login.xlsx")
            self.bookSheet = self.loadBook["Sheet1"]
        else:
            messagebox.showerror("ACE LOGIN", "File 'login.xlsx' is not in folder!")
    
    def getDataColumn(self):
        workBook = open_workbook("login.xlsx")
        workSheet = workBook.sheet_by_index(0)
        self.rowCount = workSheet.nrows
        self.rowCount += 1
        return self.rowCount

    
    def registerFunction(self, username, password):
        try:
            security = securityEncryption()
            self.bookSheet.cell(self.getDataColumn(), column=1).value = username
            self.bookSheet.cell(self.getDataColumn(), column=2).value = security.hashPassword(password)
            self.loadBook.save("login.xlsx")
            messagebox.showinfo("ACE LOGIN", "Successfully Registered!")
        except Exception:
            traceback.print_exc()
            messagebox.showerror("ACE LOGIN", "An error has occurred!")
       
    def loginFunction(self, username, password):
        try:
            security = securityEncryption()
            numRows = self.getDataColumn() + 1
            row = 1
            for rows in range(1, numRows):
                if (self.bookSheet.cell(rows, 1).value == username and self.bookSheet.cell(rows, 2).value == security.hashPassword(password)):
                    messagebox.showinfo("ACE LOGIN", "Login Found!")
                    break
                elif (rows == numRows - 1):
                    messagebox.showerror("ACE LOGIN", "Incorrect Login Info!")
                    break
        except Exception:
            traceback.print_exc()
            messagebox.showerror("ACE LOGIN", "An error has occurred!")

