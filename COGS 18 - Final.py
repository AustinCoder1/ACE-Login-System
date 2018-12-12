#!/usr/bin/env python
# coding: utf-8

# In[127]:


import hashlib
import os, sys
#from Security import *

class securityEncryption():
    def hashPassword(self, password):
        md5 = hashlib.md5()
        md5.update(bytes(password, 'utf-8'))
        return md5.hexdigest()
    
#x = securityEncryption()
#x.hashPassword("test")


# In[118]:


from xlwt import Workbook
from xlrd import open_workbook
from pathlib import Path
from tkinter import messagebox
import openpyxl
import string

class excelFunctions():
    def __init__(self):
        if (os.path.isfile("login.xlsx")):
            self.rowCount = 0
            self.columnNumber = 0
            self.loadBook = openpyxl.load_workbook("login.xlsx")
            self.bookSheet = self.loadBook["Sheet1"]
        else:
            messagebox.showerror("ACE LOGIN", "File 'login.xlsx' is not in folder!")
    
    def getDataColumn(self):
        workBook = open_workbook("login.xlsx")
        workSheet = workBook.sheet_by_index(0)
        self.rowCount = workSheet.nrows
        self.rowCount += 1
        return self.rowCount

    
    def writedata(self, username, password):
        try:
            security = securityEncryption()
            self.bookSheet.cell(self.getDataColumn(), column=1).value = username
            self.bookSheet.cell(self.getDataColumn(), column=2).value = security.hashPassword(password)
            self.loadBook.save("login.xlsx")
            messagebox.showinfo("ACE LOGIN", "Successfully Registered!")
        except Exception:
            traceback.print_exc()
            messagebox.showerror("ACE LOGIN", "An error has occurred!")
       
    def readData(self, username, password):
        try:
            numRows = self.getDataColumn() + 1
            row = 1
            print(numRows)
            for rows in range(1, numRows):
                if (self.bookSheet.cell(rows, 1).value == username and self.bookSheet.cell(rows, 2).value == password):
                    messagebox.showinfo("ACE LOGIN", "Login Found!")
                    break
                elif (rows == numRows - 1):
                    messagebox.showerror("ACE LOGIN", "Incorrect Login Info!")
                    break
        except Exception:
            traceback.print_exc()
            messagebox.showerror("ACE LOGIN", "An error has occurred!")


# In[119]:


import tkinter as tk
from tkinter import *
import os

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.master.title("|------ACE Login System------|")        
        self.widgetsLabel()
        self.widgetsButton()
        self.widgetsTextbox()
        
    def registerAccount(self):
        x = excelFunctions()
        x.writedata(self.txt_username.get(), self.txt_password.get())
        
    def loginAccount(self):
        x = excelFunctions()
        x.readData(self.txt_username.get(), self.txt_password.get())
        
    def widgetsButton(self): ##-Class that creates the buttons-##
        self.btn_login = tk.Button(root, text="Login", command=self.loginAccount)
        self.btn_login.grid(row=3, columnspan=3, sticky=E, padx=5, pady=5)
        
        self.btn_register = tk.Button(root, text="Register", command=self.registerAccount)
        self.btn_register.grid(row=3, columnspan=2, padx=5, pady=5)
        
    def widgetsTextbox(self): ##-Class that creates textboxs-##
        self.txt_username = tk.Entry(root)
        self.txt_username.grid(row=0, column=1)
        
        self.txt_password = tk.Entry(root, show="*")
        self.txt_password.grid(row=1, column=1)
        
    def widgetsLabel(self): ##-Class that creates labels-##
        
        self.lbl_username = tk.Label(root, text="Username:")
        self.lbl_username.grid(row=0, sticky=E)
        
        self.lbl_password = tk.Label(root, text="Password:")
        self.lbl_password.grid(row=1, sticky=E)
        
root = tk.Tk()
root.geometry("200x100+30+30")
createApplication = Application(master=root)
createApplication.mainloop()


# In[ ]:





# In[ ]:




